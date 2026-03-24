"""
GST ITC Reconciliation Tool — v3
==================================
New in v3:
  • Smart Python matching (Step 3) — deep invoice cleaning, amount-only matching,
    numeric token matching to catch cases fuzzy logic misses
  • Claude AI matching (Step 4) — remaining hard cases sent to Claude API
    for reasoning-based match suggestions with confidence scores

Run:  streamlit run app.py
"""

import streamlit as st
import pandas as pd
import numpy as np
import re
import json
from io import BytesIO
from datetime import datetime
from difflib import SequenceMatcher
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    import anthropic
    ANTHROPIC_AVAILABLE = True
except ImportError:
    ANTHROPIC_AVAILABLE = False

# ─────────────────────────────────────────────
# Page config
# ─────────────────────────────────────────────
st.set_page_config(page_title="GST ITC Reconciliation", page_icon="📊", layout="wide")

st.markdown("""
<style>
    .main-header {
        font-size: 2rem; font-weight: 700; color: #1a5276;
        border-bottom: 3px solid #2e86c1; padding-bottom: 10px; margin-bottom: 20px;
    }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] {
        background-color: #f0f2f6; border-radius: 6px 6px 0 0; padding: 8px 20px;
    }
    .badge-matched   { background:#d4edda; color:#155724; padding:2px 10px; border-radius:12px; font-weight:600; }
    .badge-mismatch  { background:#f8d7da; color:#721c24; padding:2px 10px; border-radius:12px; font-weight:600; }
    .badge-books     { background:#cce5ff; color:#004085; padding:2px 10px; border-radius:12px; font-weight:600; }
    .badge-portal    { background:#e2d9f3; color:#432874; padding:2px 10px; border-radius:12px; font-weight:600; }
    .badge-fuzzy     { background:#fff3cd; color:#856404; padding:2px 10px; border-radius:12px; font-weight:600; }
</style>
""", unsafe_allow_html=True)

st.markdown('<div class="main-header">📊 GST ITC Reconciliation Tool</div>', unsafe_allow_html=True)

# ─────────────────────────────────────────────
# Sidebar
# ─────────────────────────────────────────────
with st.sidebar:
    st.header("⚙️ Settings")

    with st.expander("📘 Books Column Names", expanded=False):
        b_supplier_id = st.text_input("Supplier ID / Vendor ID", value="Supplier ID", key="b_sid")
        b_gstin    = st.text_input("GSTIN",         value="GSTIN Number",          key="b_gstin")
        b_supplier = st.text_input("Supplier Name", value="Supplier",              key="b_sup")
        b_invoice  = st.text_input("Invoice No",    value="Invoice No",            key="b_inv")
        b_taxable  = st.text_input("Taxable Value", value="Sum of Taxable Value",  key="b_tax")
        b_igst     = st.text_input("IGST",          value="Sum of Integrated Tax", key="b_igst")
        b_cgst     = st.text_input("CGST",          value="Sum of Central Tax",    key="b_cgst")
        b_sgst     = st.text_input("SGST",          value="Sum of State UT Tax",   key="b_sgst")
        b_cess     = st.text_input("Cess",          value="Sum of CESS Tax",       key="b_cess")

    with st.expander("🌐 Portal Column Names", expanded=False):
        p_gstin    = st.text_input("GSTIN",         value="GSTIN of supplier",          key="p_gstin")
        p_supplier = st.text_input("Supplier Name", value="Trade/Legal name",           key="p_sup")
        p_invoice  = st.text_input("Invoice No",    value="Invoice number",             key="p_inv")
        p_taxable  = st.text_input("Taxable Value", value="Sum of Taxable Value (₹)",   key="p_tax")
        p_igst     = st.text_input("IGST",          value="Sum of Integrated Tax(₹)",   key="p_igst")
        p_cgst     = st.text_input("CGST",          value="Sum of Central Tax(₹)",      key="p_cgst")
        p_sgst     = st.text_input("SGST",          value="Sum of State/UT Tax(₹)",     key="p_sgst")
        p_cess     = st.text_input("Cess",          value="Sum of Cess(₹)",             key="p_cess")
        p_itc_avail= st.text_input("ITC Availability", value="ITC Availability",        key="p_itc")

    st.divider()
    st.subheader("🔧 Matching Settings")
    tolerance = st.number_input("Amount tolerance (₹)", min_value=0.0, value=1.0, step=0.5,
                                help="Differences within this ₹ amount are treated as matched.")
    fuzzy_threshold = st.slider("Fuzzy match sensitivity", min_value=50, max_value=100, value=80,
                                help="How similar two invoice numbers must be to be fuzzy-matched. 100 = exact only, 80 = allows minor differences.")

    st.divider()
    st.subheader("🤖 AI Matching (Step 4)")
    st.caption("Claude AI reviews remaining unmatched items and suggests matches with reasoning.")
    ai_api_key = st.text_input(
        "Anthropic API Key",
        type="password",
        placeholder="sk-ant-...",
        help="Get your key from console.anthropic.com. Key is never stored — only used for this session."
    )
    ai_model = st.selectbox(
        "AI Model",
        ["claude-haiku-4-5-20251001", "claude-sonnet-4-6"],
        index=0,
        help="Haiku is faster and cheaper (~₹0.5/run). Sonnet is more accurate (~₹3/run)."
    )
    ai_batch_size = st.slider("Max items per GSTIN sent to AI", 5, 30, 15,
        help="Per-GSTIN limit. Lower = cheaper but may miss some matches.")

# ─────────────────────────────────────────────
# Helper functions
# ─────────────────────────────────────────────
def clean_gstin(s):
    if pd.isna(s): return np.nan
    s = str(s).strip().upper()
    return np.nan if s == "" else s

def clean_invoice(s):
    if pd.isna(s): return ""
    s = str(s).strip().upper().lstrip("0")
    return s

def safe_float(col):
    return pd.to_numeric(col, errors="coerce").fillna(0.0)

def similarity(a, b):
    """Return 0-100 similarity score between two strings."""
    return round(SequenceMatcher(None, str(a), str(b)).ratio() * 100, 1)

# ─────────────────────────────────────────────
# Step 3: Smart Python matching helpers
# ─────────────────────────────────────────────
# Common invoice prefixes to strip
_INV_PREFIXES = re.compile(
    r'^(INV|BILL|PO|GRN|TXN|GST|TAX|VCH|VOUCHER|INVOICE|RCT|RCPT|RECEIPT|'
    r'SB|DM|CM|CR|DR|NF|DN|CN|MFG|MF|SL|SLV|SI|PI|LR|AWB|DC|GDN|IN|OUT)[-/\s]*',
    re.IGNORECASE
)
# Financial year codes: 2025-26 / 25-26 / 2526 / FY26 / FY2526
_YEAR_CODE = re.compile(r'(20\d{2}[-/]\d{2}|\b\d{2}[-/]\d{2}\b|FY\s*\d{2,4}|\b2[0-9]{3}\b)', re.IGNORECASE)
# Non-alphanumeric characters
_NON_ALNUM = re.compile(r'[^A-Z0-9]')

def deep_clean_invoice(s):
    """Aggressively normalise invoice number for Step 3 matching."""
    if pd.isna(s) or str(s).strip() == "": return ""
    s = str(s).strip().upper()
    s = _YEAR_CODE.sub("", s)          # strip year codes
    s = _INV_PREFIXES.sub("", s)       # strip common prefixes
    s = _NON_ALNUM.sub("", s)          # keep only letters & digits
    s = s.lstrip("0")                  # strip leading zeros
    return s.strip()

def get_numeric_tokens(s):
    """Extract all numeric sequences from an invoice string.
    E.g. 'NF/2025/00123' → ['2025', '123']"""
    return [t.lstrip("0") for t in re.findall(r'\d+', str(s)) if t.lstrip("0")]

def apply_smart_matching(only_books_df, only_portal_df, tol):
    """
    Step 3 — three Python sub-passes on items still unmatched after fuzzy:

    Pass A : Deep-clean both invoice numbers → exact match
             (strips year codes, prefixes, special chars)

    Pass B : Same GSTIN + exact total tax amount match
             (when invoice format is too different to compare, but ₹ amount is unique)

    Pass C : Numeric-token match — extract digit sequences from invoice
             and match on the longest common token
             (e.g. 'NF/25-26/000123' ↔ 'NF-123' both have core token '123')
    """
    if only_books_df.empty or only_portal_df.empty:
        return only_books_df, only_portal_df, pd.DataFrame()

    smart_rows     = []
    b_used_idx     = set()
    p_used_idx     = set()

    common_gstins = (
        set(only_books_df["GSTIN"].dropna()) &
        set(only_portal_df["GSTIN"].dropna())
    )

    for gstin in common_gstins:
        b_sub = only_books_df[only_books_df["GSTIN"] == gstin].copy()
        p_sub = only_portal_df[only_portal_df["GSTIN"] == gstin].copy()

        b_sub["_deep"] = b_sub["Invoice_No"].apply(deep_clean_invoice)
        p_sub["_deep"] = p_sub["Invoice_No"].apply(deep_clean_invoice)
        b_sub["_tokens"] = b_sub["Invoice_No"].apply(get_numeric_tokens)
        p_sub["_tokens"] = p_sub["Invoice_No"].apply(get_numeric_tokens)

        def _get_tax(row, is_books):
            return float(row.get("TotalTax_Books" if is_books else "TotalTax_Portal",
                                  row.get("Total_Tax", 0)) or 0)
        def _get_taxable(row, is_books):
            return float(row.get("Taxable_Books" if is_books else "Taxable_Portal",
                                  row.get("Taxable_Value", 0)) or 0)
        def _sid(row):
            return row.get("Supplier_ID", row.get("Supplier_ID_Books", ""))

        def _record(b_idx, p_idx, pass_name, note):
            b_row = only_books_df.loc[b_idx]
            p_row = only_portal_df.loc[p_idx]
            b_tax = _get_tax(b_row, True); p_tax = _get_tax(p_row, False)
            amt_ok = abs(b_tax - p_tax) <= tol
            status = f"Smart:{pass_name} ✓" if amt_ok else f"Smart:{pass_name} ⚠ Amt Diff"
            return {
                "Match_Type":        f"Smart — {pass_name}",
                "Supplier_ID_Books": _sid(b_row),
                "GSTIN":             gstin,
                "Supplier_Books":    b_row.get("Supplier_Name", b_row.get("Supplier_Books", "")),
                "Invoice_Books":     b_row.get("Invoice_No", ""),
                "Invoice_Portal":    p_row.get("Invoice_No", ""),
                "Note":              note,
                "Taxable_Books":     round(_get_taxable(b_row, True), 2),
                "Taxable_Portal":    round(_get_taxable(p_row, False), 2),
                "TotalTax_Books":    round(b_tax, 2),
                "TotalTax_Portal":   round(p_tax, 2),
                "Tax_Difference":    round(b_tax - p_tax, 2),
                "Status":            status,
            }

        # ── Pass A: deep-cleaned exact match ──────────────────
        p_deep_index = {row["_deep"]: idx
                        for idx, row in p_sub.iterrows()
                        if row["_deep"] and idx not in p_used_idx}

        for b_idx, b_row in b_sub.iterrows():
            if b_idx in b_used_idx: continue
            key = b_row["_deep"]
            if key and key in p_deep_index:
                p_idx = p_deep_index[key]
                if p_idx not in p_used_idx:
                    smart_rows.append(_record(b_idx, p_idx, "DeepClean",
                        f"'{b_row['Invoice_No']}' ↔ '{only_portal_df.loc[p_idx,'Invoice_No']}' after stripping year codes & prefixes"))
                    b_used_idx.add(b_idx); p_used_idx.add(p_idx)
                    del p_deep_index[key]

        # ── Pass B: exact amount match (same GSTIN, unique amount) ─
        b_rem = b_sub[~b_sub.index.isin(b_used_idx)]
        p_rem = p_sub[~p_sub.index.isin(p_used_idx)]

        # Build amount → [idx] map for portal
        p_amt_map = {}
        for p_idx, p_row in p_rem.iterrows():
            tax = round(_get_tax(only_portal_df.loc[p_idx], False), 2)
            p_amt_map.setdefault(tax, []).append(p_idx)

        for b_idx, b_row in b_rem.iterrows():
            if b_idx in b_used_idx: continue
            b_tax = round(_get_tax(only_books_df.loc[b_idx], True), 2)
            if b_tax == 0: continue           # skip zero-tax entries
            candidates = [i for i in p_amt_map.get(b_tax, []) if i not in p_used_idx]
            if len(candidates) == 1:          # unique match only
                p_idx = candidates[0]
                smart_rows.append(_record(b_idx, p_idx, "AmountMatch",
                    f"Unique exact tax amount ₹{b_tax} for this GSTIN — invoice formats differ"))
                b_used_idx.add(b_idx); p_used_idx.add(p_idx)

        # ── Pass C: numeric token match ───────────────────────
        b_rem2 = b_sub[~b_sub.index.isin(b_used_idx)]
        p_rem2 = p_sub[~p_sub.index.isin(p_used_idx)]

        # Build longest-token → [idx] map for portal
        p_token_map = {}
        for p_idx, p_row in p_rem2.iterrows():
            tokens = p_row["_tokens"]
            if tokens:
                key = max(tokens, key=len)    # use longest numeric token
                if len(key) >= 3:             # ignore very short tokens
                    p_token_map.setdefault(key, []).append(p_idx)

        for b_idx, b_row in b_rem2.iterrows():
            if b_idx in b_used_idx: continue
            tokens = b_row["_tokens"]
            if not tokens: continue
            b_key = max(tokens, key=len)
            if len(b_key) < 3: continue
            candidates = [i for i in p_token_map.get(b_key, []) if i not in p_used_idx]
            if len(candidates) == 1:
                p_idx = candidates[0]
                smart_rows.append(_record(b_idx, p_idx, "TokenMatch",
                    f"Core numeric token '{b_key}' found in both invoice numbers"))
                b_used_idx.add(b_idx); p_used_idx.add(p_idx)

    smart_df = pd.DataFrame(smart_rows)
    rem_books  = only_books_df.drop(index=list(b_used_idx)).reset_index(drop=True)
    rem_portal = only_portal_df.drop(index=list(p_used_idx)).reset_index(drop=True)
    return rem_books, rem_portal, smart_df

# ─────────────────────────────────────────────
# Step 4: Claude AI matching
# ─────────────────────────────────────────────
def apply_ai_matching(only_books_df, only_portal_df, api_key, model, batch_size, tol):
    """
    Send remaining unmatched items to Claude API grouped by GSTIN.
    Claude reasons about invoice format differences and suggests matches.
    Returns a dataframe of AI-suggested matches.
    """
    if not ANTHROPIC_AVAILABLE:
        return only_books_df, only_portal_df, pd.DataFrame(), "anthropic package not installed"

    if only_books_df.empty or only_portal_df.empty:
        return only_books_df, only_portal_df, pd.DataFrame(), "No items to match"

    client = anthropic.Anthropic(api_key=api_key)
    ai_rows    = []
    b_used_idx = set()
    p_used_idx = set()
    errors     = []

    common_gstins = (
        set(only_books_df["GSTIN"].dropna()) &
        set(only_portal_df["GSTIN"].dropna())
    )

    def _get_tax(row, is_books):
        return float(row.get("TotalTax_Books" if is_books else "TotalTax_Portal",
                              row.get("Total_Tax", 0)) or 0)
    def _get_taxable(row, is_books):
        return float(row.get("Taxable_Books" if is_books else "Taxable_Portal",
                              row.get("Taxable_Value", 0)) or 0)
    def _sid(row):
        return row.get("Supplier_ID", row.get("Supplier_ID_Books", ""))

    for gstin in common_gstins:
        b_sub = only_books_df[only_books_df["GSTIN"] == gstin].head(batch_size)
        p_sub = only_portal_df[only_portal_df["GSTIN"] == gstin].head(batch_size)
        if b_sub.empty or p_sub.empty: continue

        b_name = b_sub.iloc[0].get("Supplier_Name", b_sub.iloc[0].get("Supplier_Books", ""))

        # Build compact invoice lists for the prompt
        b_list = [{"id": str(i), "invoice": row.get("Invoice_No",""),
                   "tax": round(_get_tax(row, True), 2),
                   "taxable": round(_get_taxable(row, True), 2)}
                  for i, row in b_sub.iterrows()]
        p_list = [{"id": str(i), "invoice": row.get("Invoice_No",""),
                   "tax": round(_get_tax(row, False), 2),
                   "taxable": round(_get_taxable(row, False), 2)}
                  for i, row in p_sub.iterrows()]

        prompt = f"""You are an expert Indian Chartered Accountant doing GST ITC reconciliation.

Supplier: {b_name}  |  GSTIN: {gstin}

These invoices from BOOKS could not be matched with the PORTAL using exact or fuzzy matching:

BOOKS invoices (unmatched):
{json.dumps(b_list, indent=2)}

PORTAL invoices (unmatched):
{json.dumps(p_list, indent=2)}

Your task: Identify which Books invoice likely matches which Portal invoice.
Reasons invoices may not have matched automatically:
- Year codes like "2025-26", "FY26", "25-26" present in one but not the other
- Common prefixes like "INV-", "BILL/", "DCM/" differ between systems
- Leading zeros differ (e.g. "0123" vs "123")
- Slash/dash/dot separators differ (e.g. "NF/123" vs "NF-123")
- Amount may differ slightly due to rounding (up to ₹{tol * 2} tolerance)

Rules:
- Only match if you are reasonably confident (Medium or High confidence)
- Do NOT force-match if there is genuine uncertainty
- Use the "id" field to reference each invoice (these are row index numbers)
- Tax amounts should be close (within ₹{tol * 2})

Return ONLY a valid JSON array, no other text:
[
  {{
    "book_id": "<id from Books list>",
    "portal_id": "<id from Portal list>",
    "confidence": "High" or "Medium" or "Low",
    "reason": "<brief explanation>"
  }}
]

If no matches found, return an empty array: []"""

        try:
            response = client.messages.create(
                model=model,
                max_tokens=1024,
                messages=[{"role": "user", "content": prompt}]
            )
            raw = response.content[0].text.strip()
            # Extract JSON array from response
            json_match = re.search(r'\[.*\]', raw, re.DOTALL)
            if not json_match:
                errors.append(f"{gstin}: Could not parse AI response")
                continue
            matches = json.loads(json_match.group())

            for m in matches:
                try:
                    b_idx = int(m["book_id"])
                    p_idx = int(m["portal_id"])
                    if b_idx in b_used_idx or p_idx in p_used_idx: continue
                    if b_idx not in only_books_df.index: continue
                    if p_idx not in only_portal_df.index: continue

                    b_row = only_books_df.loc[b_idx]
                    p_row = only_portal_df.loc[p_idx]
                    b_tax = _get_tax(b_row, True)
                    p_tax = _get_tax(p_row, False)
                    confidence = m.get("confidence", "Medium")
                    amt_ok = abs(b_tax - p_tax) <= tol * 3

                    ai_rows.append({
                        "Match_Type":        f"AI — {confidence} Confidence",
                        "Supplier_ID_Books": _sid(b_row),
                        "GSTIN":             gstin,
                        "Supplier_Books":    b_row.get("Supplier_Name", b_row.get("Supplier_Books", "")),
                        "Invoice_Books":     b_row.get("Invoice_No", ""),
                        "Invoice_Portal":    p_row.get("Invoice_No", ""),
                        "AI_Confidence":     confidence,
                        "AI_Reason":         m.get("reason", ""),
                        "Taxable_Books":     round(_get_taxable(b_row, True), 2),
                        "Taxable_Portal":    round(_get_taxable(p_row, False), 2),
                        "TotalTax_Books":    round(b_tax, 2),
                        "TotalTax_Portal":   round(p_tax, 2),
                        "Tax_Difference":    round(b_tax - p_tax, 2),
                        "Status":            f"AI Matched ✓" if amt_ok else "AI Matched ⚠ Amt Diff",
                    })
                    b_used_idx.add(b_idx)
                    p_used_idx.add(p_idx)
                except (KeyError, ValueError, TypeError):
                    continue

        except Exception as e:
            errors.append(f"{gstin}: {str(e)[:80]}")
            continue

    ai_df     = pd.DataFrame(ai_rows)
    rem_books  = only_books_df.drop(index=list(b_used_idx)).reset_index(drop=True)
    rem_portal = only_portal_df.drop(index=list(p_used_idx)).reset_index(drop=True)
    err_msg    = "; ".join(errors) if errors else None
    return rem_books, rem_portal, ai_df, err_msg

def load_books(df):
    out = pd.DataFrame()
    # Supplier ID — include if column exists, else leave blank
    if b_supplier_id and b_supplier_id in df.columns:
        out["Supplier_ID"] = df[b_supplier_id].astype(str).str.strip()
    else:
        out["Supplier_ID"] = ""
    out["GSTIN"]         = df[b_gstin].apply(clean_gstin)
    out["Supplier_Name"] = df[b_supplier].astype(str).str.strip()
    out["Invoice_No"]    = df[b_invoice].apply(clean_invoice)
    out["Taxable_Value"] = safe_float(df[b_taxable])
    out["IGST"]          = safe_float(df[b_igst])
    out["CGST"]          = safe_float(df[b_cgst])
    out["SGST"]          = safe_float(df[b_sgst])
    out["Cess"]          = safe_float(df[b_cess])
    out["Total_Tax"]     = out["IGST"] + out["CGST"] + out["SGST"] + out["Cess"]
    out["Total_Value"]   = out["Taxable_Value"] + out["Total_Tax"]
    return out

def load_portal(df):
    out = pd.DataFrame()
    out["GSTIN"]         = df[p_gstin].apply(clean_gstin)
    out["Supplier_Name"] = df[p_supplier].astype(str).str.strip()
    out["Invoice_No"]    = df[p_invoice].apply(clean_invoice)
    out["Taxable_Value"] = safe_float(df[p_taxable])
    out["IGST"]          = safe_float(df[p_igst])
    out["CGST"]          = safe_float(df[p_cgst])
    out["SGST"]          = safe_float(df[p_sgst])
    out["Cess"]          = safe_float(df[p_cess])
    out["Total_Tax"]     = out["IGST"] + out["CGST"] + out["SGST"] + out["Cess"]
    out["Total_Value"]   = out["Taxable_Value"] + out["Total_Tax"]
    if p_itc_avail in df.columns:
        out["ITC_Availability"] = df[p_itc_avail].astype(str).str.strip()
    return out

# ─────────────────────────────────────────────
# Fuzzy matching helpers
# ─────────────────────────────────────────────
def _get_tax(row, books_side):
    """Safely extract total tax from a row regardless of column naming."""
    if books_side:
        return float(row.get("TotalTax_Books", row.get("Total_Tax", 0)) or 0)
    return float(row.get("TotalTax_Portal", row.get("Total_Tax", 0)) or 0)

def _get_taxable(row, books_side):
    return float(row.get("Taxable_Books", row.get("Taxable_Value", 0)) or 0)

def _build_fuzzy_row(match_type, gstin_books, gstin_portal, b_row, p_row, inv_score, amt_tol):
    b_tax  = _get_tax(b_row, True)
    p_tax  = _get_tax(p_row, False)
    b_taxv = _get_taxable(b_row, True)
    p_taxv = _get_taxable(p_row, False)
    amt_ok = abs(b_tax - p_tax) <= amt_tol

    if match_type == "same_gstin":
        status = "Fuzzy Matched ✓" if amt_ok else "Fuzzy Matched ⚠ Amt Diff"
        note   = "Same GSTIN, similar invoice no."
    else:
        status = "Probable Match ✓" if amt_ok else "Probable Match ⚠ Amt Diff"
        note   = "Diff GSTIN, similar invoice no. + amount"

    return {
        "Match_Type":        match_type.replace("_", " ").title(),
        "Supplier_ID_Books": b_row.get("Supplier_ID", b_row.get("Supplier_ID_Books", "")),
        "GSTIN_Books":       gstin_books,
        "GSTIN_Portal":      gstin_portal,
        "Supplier_Books":    b_row.get("Supplier_Name", b_row.get("Supplier_Books", "")),
        "Supplier_Portal":   p_row.get("Supplier_Name", p_row.get("Supplier_Portal", "")),
        "Invoice_Books":     b_row.get("Invoice_No", ""),
        "Invoice_Portal":    p_row.get("Invoice_No", ""),
        "Invoice_Similarity_%": inv_score,
        "Taxable_Books":     round(b_taxv, 2),
        "Taxable_Portal":    round(p_taxv, 2),
        "TotalTax_Books":    round(b_tax, 2),
        "TotalTax_Portal":   round(p_tax, 2),
        "Tax_Difference":    round(b_tax - p_tax, 2),
        "Note":              note,
        "Status":            status,
    }

def apply_fuzzy_matching(only_books_df, only_portal_df, threshold, amt_tol=1.0):
    """
    Two-pass fuzzy matching:

    Pass 1 — Same GSTIN, similar invoice number (threshold%)
              → labelled 'Fuzzy Matched'

    Pass 2 — Different GSTIN, but invoice numbers are similar (threshold%)
              AND tax amounts are close (within amt_tol * 5)
              → labelled 'Probable Match' (needs human review)
    """
    if only_books_df.empty or only_portal_df.empty:
        return only_books_df, only_portal_df, pd.DataFrame()

    fuzzy_rows      = []
    books_used_idx  = set()
    portal_used_idx = set()

    # ── Pass 1: same GSTIN fuzzy match ────────────────────────
    common_gstins = (
        set(only_books_df["GSTIN"].dropna()) &
        set(only_portal_df["GSTIN"].dropna())
    )

    for gstin in common_gstins:
        b_sub = only_books_df[only_books_df["GSTIN"] == gstin]
        p_sub = only_portal_df[only_portal_df["GSTIN"] == gstin]

        for b_idx, b_row in b_sub.iterrows():
            if b_idx in books_used_idx:
                continue
            best_score, best_p_idx = -1, None
            for p_idx, p_row in p_sub.iterrows():
                if p_idx in portal_used_idx:
                    continue
                score = similarity(b_row["Invoice_No"], p_row["Invoice_No"])
                if score >= threshold and score > best_score:
                    best_score, best_p_idx = score, p_idx

            if best_p_idx is not None:
                p_row = only_portal_df.loc[best_p_idx]
                fuzzy_rows.append(_build_fuzzy_row(
                    "same_gstin", gstin, gstin, b_row, p_row, best_score, amt_tol
                ))
                books_used_idx.add(b_idx)
                portal_used_idx.add(best_p_idx)

    # ── Pass 2: cross-GSTIN probable match ────────────────────
    # Criteria (any of the below combinations qualifies):
    #   A) Invoice similar + amounts close
    #   B) Supplier name similar + amounts close (catches name-based mismatches)
    #   C) Invoice similar + supplier name similar (even if small amount diff)
    # In all cases GSTIN must differ from pass-1 matches.

    b_rem = only_books_df.drop(index=list(books_used_idx))
    p_rem = only_portal_df.drop(index=list(portal_used_idx))

    probable_amt_tol    = max(amt_tol * 5, 10.0)   # generous tolerance for cross-GSTIN
    probable_rows       = []
    probable_b_used_idx = set()
    probable_p_used_idx = set()

    for b_idx, b_row in b_rem.iterrows():
        if b_idx in probable_b_used_idx:
            continue
        b_inv  = b_row.get("Invoice_No", "")
        b_name = b_row.get("Supplier_Name", b_row.get("Supplier_Books", ""))
        b_tax  = _get_tax(b_row, True)

        best_combined, best_p_idx = -1, None

        for p_idx, p_row in p_rem.iterrows():
            if p_idx in probable_p_used_idx:
                continue

            p_tax  = _get_tax(p_row, False)
            p_inv  = p_row.get("Invoice_No", "")
            p_name = p_row.get("Supplier_Name", p_row.get("Supplier_Portal", ""))

            amt_close = abs(b_tax - p_tax) <= probable_amt_tol

            inv_score  = similarity(b_inv,  p_inv)
            name_score = similarity(b_name, p_name)

            inv_match  = inv_score  >= threshold
            name_match = name_score >= threshold

            # Must satisfy at least one qualifying combination
            qualifies = (
                (amt_close and inv_match)   or   # A: amount + invoice
                (amt_close and name_match)  or   # B: amount + name
                (inv_match  and name_match)       # C: invoice + name (regardless of amount)
            )

            if not qualifies:
                continue

            # Combined score — weight: invoice 50%, name 30%, amount closeness 20%
            amt_score = max(0, 100 - abs(b_tax - p_tax))   # higher = closer amount
            combined  = inv_score * 0.5 + name_score * 0.3 + min(amt_score, 100) * 0.2

            if combined > best_combined:
                best_combined, best_p_idx = combined, p_idx

        if best_p_idx is not None:
            p_row   = p_rem.loc[best_p_idx]
            gstin_b = str(b_row.get("GSTIN", ""))
            gstin_p = str(p_row.get("GSTIN", ""))
            if gstin_b != gstin_p:
                p_inv  = p_row.get("Invoice_No", "")
                p_name = p_row.get("Supplier_Name", p_row.get("Supplier_Portal", ""))
                inv_s  = similarity(b_inv, p_inv)
                # Store best invoice score for display
                row = _build_fuzzy_row(
                    "cross_gstin", gstin_b, gstin_p, b_row, p_row, inv_s, amt_tol
                )
                row["Name_Similarity_%"] = similarity(b_name, p_name)
                row["Combined_Score_%"]  = round(best_combined, 1)
                probable_rows.append(row)
                probable_b_used_idx.add(b_idx)
                probable_p_used_idx.add(best_p_idx)

    all_fuzzy = pd.DataFrame(fuzzy_rows + probable_rows)

    all_used_b = books_used_idx  | probable_b_used_idx
    all_used_p = portal_used_idx | probable_p_used_idx

    remaining_books  = only_books_df.drop(index=list(all_used_b))
    remaining_portal = only_portal_df.drop(index=list(all_used_p))

    return (
        remaining_books.reset_index(drop=True),
        remaining_portal.reset_index(drop=True),
        all_fuzzy
    )

# ─────────────────────────────────────────────
# GSTIN-Level Reconciliation
# ─────────────────────────────────────────────
def reconcile_gstin_level(books_df, portal_df, tol):
    b_agg = books_df.dropna(subset=["GSTIN"]).groupby("GSTIN", as_index=False).agg(
        Supplier_ID_Books    =("Supplier_ID",    "first"),
        Supplier_Name_Books  =("Supplier_Name",  "first"),
        Invoice_Count_Books  =("Invoice_No",     "count"),
        Taxable_Books        =("Taxable_Value",  "sum"),
        IGST_Books           =("IGST",           "sum"),
        CGST_Books           =("CGST",           "sum"),
        SGST_Books           =("SGST",           "sum"),
        Cess_Books           =("Cess",           "sum"),
        TotalTax_Books       =("Total_Tax",      "sum"),
    )
    p_agg = portal_df.dropna(subset=["GSTIN"]).groupby("GSTIN", as_index=False).agg(
        Supplier_Name_Portal =("Supplier_Name",  "first"),
        Invoice_Count_Portal =("Invoice_No",     "count"),
        Taxable_Portal       =("Taxable_Value",  "sum"),
        IGST_Portal          =("IGST",           "sum"),
        CGST_Portal          =("CGST",           "sum"),
        SGST_Portal          =("SGST",           "sum"),
        Cess_Portal          =("Cess",           "sum"),
        TotalTax_Portal      =("Total_Tax",      "sum"),
    )

    merged = pd.merge(b_agg, p_agg, on="GSTIN", how="outer", indicator=True)

    for s in ["Taxable","IGST","CGST","SGST","Cess","TotalTax"]:
        merged[f"{s}_Books"]  = merged[f"{s}_Books"].fillna(0)
        merged[f"{s}_Portal"] = merged[f"{s}_Portal"].fillna(0)
        merged[f"Diff_{s}"]   = merged[f"{s}_Books"] - merged[f"{s}_Portal"]

    merged["Invoice_Count_Books"]  = merged["Invoice_Count_Books"].fillna(0).astype(int)
    merged["Invoice_Count_Portal"] = merged["Invoice_Count_Portal"].fillna(0).astype(int)

    def classify(row):
        if row["_merge"] == "left_only":  return "Only in Books"
        if row["_merge"] == "right_only": return "Only in Portal"
        # Taxable value difference is IGNORED — only total tax decides match/mismatch
        if abs(row["Diff_TotalTax"]) <= tol:
            return "Matched"
        return "Mismatched"

    def get_remarks(row):
        if row["Status"] not in ("Matched", "Mismatched"):
            return ""
        remarks = []
        # Taxable difference note (informational — doesn't cause mismatch)
        if abs(row.get("Diff_Taxable", 0)) > tol:
            remarks.append(f"ℹ️ Taxable differs by ₹{abs(row['Diff_Taxable']):.2f} — tax matches")
        # IGST ↔ CGST+SGST swap detection
        igst_b = float(row.get("IGST_Books",  0) or 0)
        cgst_b = float(row.get("CGST_Books",  0) or 0)
        sgst_b = float(row.get("SGST_Books",  0) or 0)
        igst_p = float(row.get("IGST_Portal", 0) or 0)
        cgst_p = float(row.get("CGST_Portal", 0) or 0)
        sgst_p = float(row.get("SGST_Portal", 0) or 0)
        if abs(igst_b - (cgst_p + sgst_p)) <= tol and igst_b > tol:
            remarks.append("⚠️ Tax head swap: IGST in Books = CGST+SGST in Portal — verify inter/intra-state")
        elif abs((cgst_b + sgst_b) - igst_p) <= tol and igst_p > tol:
            remarks.append("⚠️ Tax head swap: CGST+SGST in Books = IGST in Portal — verify inter/intra-state")
        return " | ".join(remarks)

    merged["Status"]  = merged.apply(classify, axis=1)
    merged["Remarks"] = merged.apply(get_remarks, axis=1)
    merged.drop(columns=["_merge"], inplace=True)
    num_cols = merged.select_dtypes(include=[np.number]).columns
    merged[num_cols] = merged[num_cols].round(2)

    matched     = merged[merged["Status"] == "Matched"].copy()
    mismatched  = merged[merged["Status"] == "Mismatched"].copy()
    only_books  = merged[merged["Status"] == "Only in Books"].copy()
    only_portal = merged[merged["Status"] == "Only in Portal"].copy()
    no_gstin    = books_df[books_df["GSTIN"].isna()].copy()

    return merged, matched, mismatched, only_books, only_portal, no_gstin

# ─────────────────────────────────────────────
# Invoice-Level Reconciliation
# ─────────────────────────────────────────────
def reconcile_invoice_level(books_df, portal_df, tol):
    b = books_df.dropna(subset=["GSTIN"]).copy()
    p = portal_df.dropna(subset=["GSTIN"]).copy()

    b["_key"] = b["GSTIN"] + "|" + b["Invoice_No"]
    p["_key"] = p["GSTIN"] + "|" + p["Invoice_No"]

    b_r = b.rename(columns={"Supplier_ID":"Supplier_ID_Books","Supplier_Name":"Supplier_Books",
                             "Taxable_Value":"Taxable_Books",
                             "IGST":"IGST_Books","CGST":"CGST_Books","SGST":"SGST_Books",
                             "Cess":"Cess_Books","Total_Tax":"TotalTax_Books","Total_Value":"TotalValue_Books"})
    p_r = p.rename(columns={"Supplier_Name":"Supplier_Portal","Taxable_Value":"Taxable_Portal",
                             "IGST":"IGST_Portal","CGST":"CGST_Portal","SGST":"SGST_Portal",
                             "Cess":"Cess_Portal","Total_Tax":"TotalTax_Portal","Total_Value":"TotalValue_Portal"})

    b_cols = ["_key","GSTIN","Invoice_No","Supplier_ID_Books","Supplier_Books",
              "Taxable_Books","IGST_Books","CGST_Books","SGST_Books","Cess_Books","TotalTax_Books","TotalValue_Books"]
    p_cols = ["_key","Supplier_Portal","Taxable_Portal","IGST_Portal","CGST_Portal",
              "SGST_Portal","Cess_Portal","TotalTax_Portal","TotalValue_Portal"]
    if "ITC_Availability" in p_r.columns:
        p_cols.append("ITC_Availability")

    merged = pd.merge(b_r[b_cols], p_r[p_cols], on="_key", how="outer", indicator=True)

    # Fill GSTIN/Invoice for portal-only rows
    for idx in merged[merged["_merge"] == "right_only"].index:
        key = merged.loc[idx, "_key"]
        if pd.notna(key) and "|" in str(key):
            parts = str(key).split("|", 1)
            merged.loc[idx, "GSTIN"]      = parts[0]
            merged.loc[idx, "Invoice_No"] = parts[1]

    for s in ["Taxable","IGST","CGST","SGST","Cess","TotalTax"]:
        merged[f"{s}_Books"]  = merged[f"{s}_Books"].fillna(0)
        merged[f"{s}_Portal"] = merged[f"{s}_Portal"].fillna(0)
        merged[f"Diff_{s}"]   = merged[f"{s}_Books"] - merged[f"{s}_Portal"]

    def classify(row):
        if row["_merge"] == "left_only":  return "Only in Books"
        if row["_merge"] == "right_only": return "Only in Portal"
        # Taxable value difference is IGNORED — only total tax decides match/mismatch
        if abs(row["Diff_TotalTax"]) <= tol:
            return "Matched"
        return "Mismatched"

    def get_remarks(row):
        if row["Status"] not in ("Matched", "Mismatched"):
            return ""
        remarks = []
        if abs(row.get("Diff_Taxable", 0)) > tol:
            remarks.append(f"ℹ️ Taxable differs by ₹{abs(row['Diff_Taxable']):.2f} — tax matches")
        igst_b = float(row.get("IGST_Books",  0) or 0)
        cgst_b = float(row.get("CGST_Books",  0) or 0)
        sgst_b = float(row.get("SGST_Books",  0) or 0)
        igst_p = float(row.get("IGST_Portal", 0) or 0)
        cgst_p = float(row.get("CGST_Portal", 0) or 0)
        sgst_p = float(row.get("SGST_Portal", 0) or 0)
        if abs(igst_b - (cgst_p + sgst_p)) <= tol and igst_b > tol:
            remarks.append("⚠️ Tax head swap: IGST in Books = CGST+SGST in Portal — verify inter/intra-state")
        elif abs((cgst_b + sgst_b) - igst_p) <= tol and igst_p > tol:
            remarks.append("⚠️ Tax head swap: CGST+SGST in Books = IGST in Portal — verify inter/intra-state")
        return " | ".join(remarks)

    merged["Status"]  = merged.apply(classify, axis=1)
    merged["Remarks"] = merged.apply(get_remarks, axis=1)
    merged.drop(columns=["_merge","_key"], inplace=True)
    num_cols = merged.select_dtypes(include=[np.number]).columns
    merged[num_cols] = merged[num_cols].round(2)

    matched     = merged[merged["Status"] == "Matched"].copy()
    mismatched  = merged[merged["Status"] == "Mismatched"].copy()
    only_books  = merged[merged["Status"] == "Only in Books"].copy()
    only_portal = merged[merged["Status"] == "Only in Portal"].copy()
    no_gstin    = books_df[books_df["GSTIN"].isna()].copy()

    return merged, matched, mismatched, only_books, only_portal, no_gstin

# ─────────────────────────────────────────────
# Colour-coded Excel export
# ─────────────────────────────────────────────
FILL_GREEN  = PatternFill("solid", fgColor="C6EFCE")   # Matched
FILL_RED    = PatternFill("solid", fgColor="FFC7CE")   # Mismatched
FILL_BLUE   = PatternFill("solid", fgColor="BDD7EE")   # Only in Books
FILL_PURPLE = PatternFill("solid", fgColor="E2CFEF")   # Only in Portal
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")   # Fuzzy Matched
FILL_HEADER = PatternFill("solid", fgColor="1F4E79")   # Header row
FONT_HEADER = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
FONT_BOLD   = Font(bold=True, name="Calibri", size=10)
FONT_NORMAL = Font(name="Calibri", size=10)

STATUS_FILL = {
    "Matched":        FILL_GREEN,
    "Mismatched":     FILL_RED,
    "Only in Books":  FILL_BLUE,
    "Only in Portal": FILL_PURPLE,
}

def style_sheet(ws, df, status_col="Status"):
    """Apply colour-coding to a worksheet based on Status column."""
    # Header row
    for cell in ws[1]:
        cell.fill   = FILL_HEADER
        cell.font   = FONT_HEADER
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Data rows
    status_idx = None
    if status_col in df.columns:
        status_idx = list(df.columns).index(status_col) + 1  # 1-based

    for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
        status_val = ""
        if status_idx:
            status_val = ws.cell(row=row_idx, column=status_idx).value or ""

        # Determine fill for this row
        fill = None
        for key, f in STATUS_FILL.items():
            if key in str(status_val):
                fill = f
                break
        if fill is None and "Fuzzy" in str(status_val):
            fill = FILL_YELLOW

        for cell in row:
            cell.font      = FONT_NORMAL
            cell.alignment = Alignment(vertical="center")
            if fill:
                cell.fill = fill

    # Auto-width columns
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 4, 40)

    # Freeze header row
    ws.freeze_panes = "A2"

def build_summary_sheet(ws, stats, amt_df):
    """Write a nicely formatted Summary sheet."""
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 20

    # Title
    ws["A1"] = "GST ITC RECONCILIATION REPORT"
    ws["A1"].font = Font(bold=True, size=14, color="1F4E79", name="Calibri")
    ws["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y  %H:%M')}"
    ws["A2"].font = Font(italic=True, size=10, color="595959", name="Calibri")

    # Count stats
    ws["A4"] = "RECORD COUNT SUMMARY"
    ws["A4"].font = FONT_BOLD
    headers = ["Category", "Count"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")

    stat_rows = [
        ("✅ Matched",        stats.get("matched", 0),        FILL_GREEN),
        ("⚠️ Mismatched",     stats.get("mismatched", 0),     FILL_RED),
        ("🔀 Fuzzy Matched",  stats.get("fuzzy", 0),          FILL_YELLOW),
        ("📘 Only in Books",  stats.get("only_books", 0),     FILL_BLUE),
        ("🌐 Only in Portal", stats.get("only_portal", 0),    FILL_PURPLE),
        ("🚫 No GSTIN",       stats.get("no_gstin", 0),       PatternFill("solid", fgColor="EEEEEE")),
    ]
    for i, (label, count, fill) in enumerate(stat_rows, start=6):
        ws.cell(row=i, column=1, value=label).fill = fill
        ws.cell(row=i, column=2, value=count).fill = fill
        ws.cell(row=i, column=1).font = FONT_NORMAL
        ws.cell(row=i, column=2).font = FONT_NORMAL

    # Amount summary
    ws["A13"] = "AMOUNT SUMMARY (₹)"
    ws["A13"].font = FONT_BOLD
    for c, h in enumerate(amt_df.columns, 1):
        cell = ws.cell(row=14, column=c, value=h)
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")
    for r, row in enumerate(amt_df.itertuples(index=False), start=15):
        for c, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = FONT_NORMAL
            if c > 1:
                cell.number_format = "#,##0.00"

    # Colour legend
    ws["A22"] = "COLOUR LEGEND"
    ws["A22"].font = FONT_BOLD
    legend = [
        ("Green",  "Matched — amounts agree within tolerance",       FILL_GREEN),
        ("Red",    "Mismatched — found in both but amounts differ",   FILL_RED),
        ("Yellow", "Fuzzy Matched — invoice numbers nearly match",    FILL_YELLOW),
        ("Blue",   "Only in Books — not found on portal",             FILL_BLUE),
        ("Purple", "Only in Portal — not found in books",             FILL_PURPLE),
    ]
    for i, (colour, desc, fill) in enumerate(legend, start=23):
        ws.cell(row=i, column=1, value=colour).fill = fill
        ws.cell(row=i, column=2, value=desc).fill   = fill
        ws.cell(row=i, column=1).font = FONT_NORMAL
        ws.cell(row=i, column=2).font = FONT_NORMAL

def to_coloured_excel(sheets_dict, stats, amt_df):
    """Build a colour-coded Excel workbook."""
    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        # Write all dataframes first
        for name, df in sheets_dict.items():
            if df is None or len(df) == 0:
                df = pd.DataFrame({"Info": ["No records in this category"]})
            df.to_excel(writer, sheet_name=name[:31], index=False)

        wb = writer.book

        # Style each data sheet
        for name, df in sheets_dict.items():
            ws = wb[name[:31]]
            if df is not None and len(df) > 0:
                style_sheet(ws, df)

        # Build summary sheet
        if "Summary" in wb.sheetnames:
            wb.remove(wb["Summary"])
        ws_sum = wb.create_sheet("Summary", 0)
        build_summary_sheet(ws_sum, stats, amt_df)

    buf.seek(0)
    return buf

# ─────────────────────────────────────────────
# Main UI
# ─────────────────────────────────────────────
col1, col2 = st.columns(2)
with col1:
    books_file  = st.file_uploader("📘 Upload **Books** file (.xlsx)", type=["xlsx","xls"])
with col2:
    portal_file = st.file_uploader("🌐 Upload **Portal** file (.xlsx)", type=["xlsx","xls"])

recon_mode = st.radio(
    "🔍 Reconciliation Mode",
    ["GSTIN-Level Summary", "GSTIN + Invoice No (Detailed)"],
    horizontal=True,
    help="Summary mode aggregates per GSTIN. Detailed mode matches each invoice line."
)

run_btn = st.button("🚀 Run Reconciliation", type="primary", use_container_width=True)

# ─────────────────────────────────────────────
if run_btn:
    if not books_file or not portal_file:
        st.error("Please upload both Books and Portal files.")
        st.stop()

    with st.spinner("Reading files..."):
        try:
            raw_books  = pd.read_excel(books_file)
            raw_portal = pd.read_excel(portal_file)
        except Exception as e:
            st.error(f"Error reading files: {e}")
            st.stop()

    missing_b = [c for c in [b_gstin,b_supplier,b_invoice,b_taxable,b_igst,b_cgst,b_sgst,b_cess]
                 if c not in raw_books.columns]
    missing_p = [c for c in [p_gstin,p_supplier,p_invoice,p_taxable,p_igst,p_cgst,p_sgst,p_cess]
                 if c not in raw_portal.columns]
    if missing_b:
        st.error(f"Books file missing columns: {missing_b}. Check sidebar → Books Column Names.")
        st.stop()
    if missing_p:
        st.error(f"Portal file missing columns: {missing_p}. Check sidebar → Portal Column Names.")
        st.stop()

    with st.spinner("Normalising data..."):
        books_df  = load_books(raw_books)
        portal_df = load_portal(raw_portal)

    with st.spinner("Running exact reconciliation..."):
        if recon_mode == "GSTIN-Level Summary":
            full, matched, mismatched, only_books, only_portal, no_gstin = \
                reconcile_gstin_level(books_df, portal_df, tolerance)
        else:
            full, matched, mismatched, only_books, only_portal, no_gstin = \
                reconcile_invoice_level(books_df, portal_df, tolerance)

    with st.spinner("Running fuzzy matching on unmatched invoices..."):
        only_books_rem, only_portal_rem, fuzzy_df = apply_fuzzy_matching(
            only_books.copy(), only_portal.copy(), fuzzy_threshold, amt_tol=tolerance
        )

    with st.spinner("Running smart Python matching (deep clean, amount, token)..."):
        only_books_rem, only_portal_rem, smart_df = apply_smart_matching(
            only_books_rem, only_portal_rem, tolerance
        )

    # Store in session state so AI can use them later without re-running
    st.session_state["only_books_rem"]  = only_books_rem
    st.session_state["only_portal_rem"] = only_portal_rem
    st.session_state["matched"]         = matched
    st.session_state["mismatched"]      = mismatched
    st.session_state["fuzzy_df"]        = fuzzy_df
    st.session_state["smart_df"]        = smart_df
    st.session_state["no_gstin"]        = no_gstin
    st.session_state["full"]            = full
    st.session_state["books_df"]        = books_df
    st.session_state["portal_df"]       = portal_df
    st.session_state["recon_done"]      = True
    st.session_state["ai_df"]           = pd.DataFrame()   # reset AI results on new run

    # ── Amount summary ──
    bk = books_df.dropna(subset=["GSTIN"])
    amt_rows = []
    for label, bc, pc in [
        ("Taxable Value","Taxable_Value","Taxable_Value"),
        ("IGST","IGST","IGST"), ("CGST","CGST","CGST"), ("SGST","SGST","SGST"),
        ("Cess","Cess","Cess"), ("Total Tax","Total_Tax","Total_Tax"),
    ]:
        b_sum = round(bk[bc].sum(), 2)
        p_sum = round(portal_df[pc].sum(), 2)
        amt_rows.append({"Head": label, "Books (₹)": b_sum, "Portal (₹)": p_sum, "Difference (₹)": round(b_sum - p_sum, 2)})
    amt_df = pd.DataFrame(amt_rows)

    stats = {
        "matched":    len(matched),
        "mismatched": len(mismatched),
        "fuzzy":      len(fuzzy_df),
        "smart":      len(smart_df),
        "only_books": len(only_books_rem),
        "only_portal":len(only_portal_rem),
        "no_gstin":   len(no_gstin),
    }

    # ── Summary metrics ──
    st.divider()
    st.subheader("📊 Reconciliation Summary")
    c1, c2, c3, c4, c5, c6, c7 = st.columns(7)
    c1.metric("✅ Matched",        len(matched))
    c2.metric("⚠️ Mismatched",     len(mismatched))
    c3.metric("🔀 Fuzzy Matched",  len(fuzzy_df))
    c4.metric("🧠 Smart Matched",  len(smart_df))
    c5.metric("📘 Only in Books",  len(only_books_rem))
    c6.metric("🌐 Only in Portal", len(only_portal_rem))
    c7.metric("🚫 No GSTIN",       len(no_gstin))

    total_near = len(fuzzy_df) + len(smart_df)
    if total_near > 0:
        st.info(f"🔀 Fuzzy + Smart matching found **{total_near} additional near-matches** "
                f"beyond exact matching. Review the Fuzzy and Smart Matched tabs.")

    # ── Amount summary ──
    st.subheader("💰 Amount Summary")
    def colour_diff(val):
        if isinstance(val, (int, float)):
            if val > 1:   return "color: red; font-weight: bold"
            if val < -1:  return "color: red; font-weight: bold"
        return ""
    st.dataframe(amt_df.style.applymap(colour_diff, subset=["Difference (₹)"]),
                 use_container_width=True, hide_index=True)

    # ── Detailed tabs ──
    st.divider()
    tab1, tab2, tab3, tab4, tab5, tab6, tab7 = st.tabs([
        f"✅ Matched ({len(matched)})",
        f"⚠️ Mismatched ({len(mismatched)})",
        f"🔀 Fuzzy ({len(fuzzy_df)})",
        f"🧠 Smart ({len(smart_df)})",
        f"📘 Only Books ({len(only_books_rem)})",
        f"🌐 Only Portal ({len(only_portal_rem)})",
        f"🚫 No GSTIN ({len(no_gstin)})",
    ])

    with tab1:
        st.caption("Exact matches — found in both Books and Portal with matching amounts.")
        if len(matched) > 0: st.dataframe(matched, use_container_width=True, hide_index=True)
        else: st.success("No exact matches found.")

    with tab2:
        st.caption("Found in both Books and Portal but amounts do not match — needs investigation.")
        if len(mismatched) > 0: st.dataframe(mismatched, use_container_width=True, hide_index=True)
        else: st.success("No mismatches — everything balances!")

    with tab3:
        st.caption(
            f"**Fuzzy Matched** = same GSTIN, invoice strings ≥{fuzzy_threshold}% similar.  "
            f"**Probable Match** = different GSTIN, similar invoice + amount — likely a GSTIN typo. "
            f"Review all manually before accepting."
        )
        if len(fuzzy_df) > 0:
            same_gstin  = fuzzy_df[~fuzzy_df["Match_Type"].str.contains("Cross", na=False)] if "Match_Type" in fuzzy_df.columns else fuzzy_df
            cross_gstin = fuzzy_df[fuzzy_df["Match_Type"].str.contains("Cross", na=False)]  if "Match_Type" in fuzzy_df.columns else pd.DataFrame()
            if len(same_gstin) > 0:
                st.markdown("**🔀 Fuzzy Matched — Same GSTIN, similar invoice number:**")
                st.dataframe(same_gstin, use_container_width=True, hide_index=True)
            if len(cross_gstin) > 0:
                st.markdown("**🟡 Probable Match — Different GSTIN:**")
                st.warning("⚠️ GSTIN mismatch — verify before using for ITC claim.")
                st.dataframe(cross_gstin, use_container_width=True, hide_index=True)
        else:
            st.info("No fuzzy matches found. Try lowering the sensitivity slider in the sidebar.")

    with tab4:
        st.caption(
            "Smart Python matching found these after fuzzy matching.  "
            "**DeepClean** = matched after stripping year codes & prefixes.  "
            "**AmountMatch** = unique exact tax amount for this GSTIN.  "
            "**TokenMatch** = matched on core numeric sequence in the invoice number."
        )
        if len(smart_df) > 0:
            for pass_name, label in [("DeepClean","🧹 Deep Clean — stripped year codes & prefixes"),
                                      ("AmountMatch","💰 Amount Match — unique tax amount"),
                                      ("TokenMatch","🔢 Token Match — matched on core invoice number")]:
                sub = smart_df[smart_df["Match_Type"].str.contains(pass_name, na=False)] if "Match_Type" in smart_df.columns else pd.DataFrame()
                if len(sub) > 0:
                    st.markdown(f"**{label} ({len(sub)} records):**")
                    st.dataframe(sub, use_container_width=True, hide_index=True)
        else:
            st.info("No additional smart matches found.")

    with tab5:
        st.caption("In your Books but NOT found on GST Portal even after all matching passes.")
        if len(only_books_rem) > 0: st.dataframe(only_books_rem, use_container_width=True, hide_index=True)
        else: st.info("No unmatched entries in Books.")

    with tab6:
        st.caption("On GST Portal but NOT found in Books even after all matching passes.")
        if len(only_portal_rem) > 0: st.dataframe(only_portal_rem, use_container_width=True, hide_index=True)
        else: st.info("No unmatched entries in Portal.")

    with tab7:
        st.caption("Books entries with no GSTIN (Petty Cash, unregistered vendors).")
        if len(no_gstin) > 0: st.dataframe(no_gstin, use_container_width=True, hide_index=True)
        else: st.info("All books entries have a GSTIN.")

    # ──────────────────────────────────────────
    # AI MATCHING SECTION
    # ──────────────────────────────────────────
    st.divider()
    st.subheader("🤖 Step 4: Claude AI Matching")

    remaining_b_count = len(only_books_rem)
    remaining_p_count = len(only_portal_rem)

    if remaining_b_count == 0 and remaining_p_count == 0:
        st.success("✅ Nothing left for AI to review — all items matched by Python!")
    else:
        st.info(
            f"**{remaining_b_count}** books entries and **{remaining_p_count}** portal entries "
            f"still unmatched. Claude AI will review these and suggest matches with reasoning."
        )
        if not ai_api_key:
            st.warning("🔑 Enter your Anthropic API key in the sidebar to enable AI matching.")
        elif not ANTHROPIC_AVAILABLE:
            st.error("⚠️ The `anthropic` package is not installed. Add it to requirements.txt and redeploy.")
        else:
            ai_btn = st.button("🤖 Run AI Matching on Remaining Items", type="primary", use_container_width=True)
            if ai_btn:
                with st.spinner("Claude AI is analysing remaining unmatched items... (may take 30-60 sec)"):
                    rem_b = st.session_state.get("only_books_rem",  only_books_rem)
                    rem_p = st.session_state.get("only_portal_rem", only_portal_rem)
                    rem_b2, rem_p2, ai_df_result, err = apply_ai_matching(
                        rem_b, rem_p, ai_api_key, ai_model, ai_batch_size, tolerance
                    )
                    st.session_state["ai_df"]           = ai_df_result
                    st.session_state["only_books_rem"]  = rem_b2
                    st.session_state["only_portal_rem"] = rem_p2
                    if err:
                        st.warning(f"Some GSTINs had errors: {err}")

        # Show AI results if available
        ai_df = st.session_state.get("ai_df", pd.DataFrame())
        if len(ai_df) > 0:
            st.markdown(f"### 🤖 AI Suggested Matches — {len(ai_df)} found")
            st.caption(
                "These matches were identified by Claude AI. "
                "**High confidence** = very likely correct. "
                "**Medium confidence** = probable, verify manually. "
                "Review the AI_Reason column to understand why each match was suggested."
            )
            high   = ai_df[ai_df["AI_Confidence"] == "High"]   if "AI_Confidence" in ai_df.columns else pd.DataFrame()
            medium = ai_df[ai_df["AI_Confidence"] == "Medium"] if "AI_Confidence" in ai_df.columns else pd.DataFrame()
            low    = ai_df[ai_df["AI_Confidence"] == "Low"]    if "AI_Confidence" in ai_df.columns else pd.DataFrame()

            if len(high) > 0:
                st.markdown(f"**✅ High Confidence ({len(high)}):**")
                st.dataframe(high, use_container_width=True, hide_index=True)
            if len(medium) > 0:
                st.markdown(f"**🟡 Medium Confidence ({len(medium)}) — verify before accepting:**")
                st.dataframe(medium, use_container_width=True, hide_index=True)
            if len(low) > 0:
                st.markdown(f"**🟠 Low Confidence ({len(low)}) — treat as suggestions only:**")
                st.dataframe(low, use_container_width=True, hide_index=True)

    # ──────────────────────────────────────────
    # SUPPLIER DRILL-DOWN
    # ──────────────────────────────────────────
    st.divider()
    st.subheader("🔎 Supplier Drill-Down")
    st.caption("Pick any GSTIN to see all its invoices from both Books and Portal side by side.")

    all_gstins = sorted(
        set(books_df["GSTIN"].dropna().tolist()) | set(portal_df["GSTIN"].dropna().tolist())
    )

    # Build a label like "27AAJCB2354C1ZZ — SUPPLIER NAME"
    gstin_labels = {}
    for g in all_gstins:
        b_name = books_df[books_df["GSTIN"] == g]["Supplier_Name"].values
        p_name = portal_df[portal_df["GSTIN"] == g]["Supplier_Name"].values
        name = b_name[0] if len(b_name) > 0 else (p_name[0] if len(p_name) > 0 else "")
        gstin_labels[g] = f"{g}  —  {name}"

    selected_gstin = st.selectbox(
        "Select Supplier (GSTIN)",
        options=["— Select a GSTIN —"] + all_gstins,
        format_func=lambda g: gstin_labels.get(g, g) if g != "— Select a GSTIN —" else g
    )

    if selected_gstin != "— Select a GSTIN —":
        b_inv = books_df[books_df["GSTIN"] == selected_gstin].copy()
        p_inv = portal_df[portal_df["GSTIN"] == selected_gstin].copy()

        dc1, dc2 = st.columns(2)

        with dc1:
            st.markdown(f"**📘 Books** — {len(b_inv)} invoice(s)")
            if len(b_inv) > 0:
                show_b = b_inv[["Invoice_No","Taxable_Value","IGST","CGST","SGST","Cess","Total_Tax"]].reset_index(drop=True)
                st.dataframe(show_b, use_container_width=True, hide_index=True)
                st.markdown(f"**Total Tax (Books): ₹{b_inv['Total_Tax'].sum():,.2f}**")
            else:
                st.info("No invoices in Books for this GSTIN.")

        with dc2:
            st.markdown(f"**🌐 Portal** — {len(p_inv)} invoice(s)")
            if len(p_inv) > 0:
                p_show_cols = ["Invoice_No","Taxable_Value","IGST","CGST","SGST","Cess","Total_Tax"]
                if "ITC_Availability" in p_inv.columns:
                    p_show_cols.append("ITC_Availability")
                show_p = p_inv[p_show_cols].reset_index(drop=True)
                st.dataframe(show_p, use_container_width=True, hide_index=True)
                st.markdown(f"**Total Tax (Portal): ₹{p_inv['Total_Tax'].sum():,.2f}**")
            else:
                st.info("No invoices on Portal for this GSTIN.")

        # Difference callout
        b_total = b_inv["Total_Tax"].sum()
        p_total = p_inv["Total_Tax"].sum()
        diff    = round(b_total - p_total, 2)
        if abs(diff) <= tolerance:
            st.success(f"✅ Totals match for this supplier. Difference: ₹{diff:,.2f}")
        else:
            st.error(f"⚠️ Tax difference for this supplier: ₹{diff:,.2f}  "
                     f"(Books ₹{b_total:,.2f} vs Portal ₹{p_total:,.2f})")

    # ──────────────────────────────────────────
    # DOWNLOAD — Colour-coded Excel
    # ──────────────────────────────────────────
    st.divider()
    st.subheader("📥 Download Colour-Coded Report")

    ai_df_export    = st.session_state.get("ai_df", pd.DataFrame())
    final_books_rem = st.session_state.get("only_books_rem",  only_books_rem)
    final_portal_rem= st.session_state.get("only_portal_rem", only_portal_rem)

    sheets = {
        "Matched":          matched,
        "Mismatched":       mismatched,
        "Fuzzy Matched":    fuzzy_df.copy() if len(fuzzy_df) > 0 else pd.DataFrame(),
        "Smart Matched":    smart_df.copy() if len(smart_df) > 0 else pd.DataFrame(),
        "AI Matched":       ai_df_export.copy() if len(ai_df_export) > 0 else pd.DataFrame(),
        "Only in Books":    final_books_rem,
        "Only in Portal":   final_portal_rem,
        "No GSTIN":         no_gstin,
        "Full Recon":       full,
    }

    # Update stats with AI results
    stats["smart"] = len(smart_df)
    stats["ai"]    = len(ai_df_export)

    timestamp  = datetime.now().strftime("%Y%m%d_%H%M%S")
    mode_tag   = "GSTIN" if recon_mode == "GSTIN-Level Summary" else "Invoice"
    filename   = f"GST_Reconciliation_{mode_tag}_{timestamp}.xlsx"

    excel_bytes = to_coloured_excel(sheets, stats, amt_df)

    st.download_button(
        label="⬇️ Download Full Colour-Coded Report",
        data=excel_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        type="primary",
        use_container_width=True,
    )

# ── Footer ──
st.divider()
st.caption("GST ITC Reconciliation Tool v3 • Exact → Fuzzy → Smart Python → Claude AI Matching • Colour-Coded Export • Supplier Drill-Down • Built for CA Professionals")
